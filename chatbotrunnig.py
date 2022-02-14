from PyQt5.QtGui import QColor
from keras.models import load_model
import nltk
from snowballstemmer import TurkishStemmer
import json
import numpy
import random
from openpyxl import load_workbook
import numpy as np
import sys
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *

global kategorilist, puanlist, kontrol
kontrol = False
kategorilist = []
puanlist = []

imdblistesi = []
wb = load_workbook("imdb.xlsx")
ws = wb.active
for i in range(1, 101):
    imdblistesi.append([ws.cell(i, 1).value, int(ws.cell(i, 2).value), float(ws.cell(i, 3).value), ws.cell(i, 4).value,
                        ws.cell(i, 5).value])

with open(r"data.json",encoding='utf-8') as file:
    data = json.load(file)

stemmer = TurkishStemmer()
model = load_model('my_model.h5')
words = []
labels = []
docs_x = []
docs_y = []

for intent in data["intents"]:
    for pattern in intent["patterns"]:
        wrds = nltk.word_tokenize(pattern)
        words.extend(wrds)
        docs_x.append(wrds)
        docs_y.append(intent["tag"])

    if intent["tag"] not in labels:
        labels.append(intent["tag"])

words = [stemmer.stemWord(w.lower()) for w in words if w != "?"]
words = sorted(list(set(words)))

labels = sorted(labels)

training = []
output = []
out_empty = [0 for _ in range(len(labels))]

for x, doc in enumerate(docs_x):
    bag = []

    wrds = [stemmer.stemWord(w.lower()) for w in doc]

    for w in words:
        if w in wrds:
            bag.append(1)
        else:
            bag.append(0)

    output_row = out_empty[:]
    output_row[labels.index(docs_y[x])] = 1

    training.append(bag)
    output.append(output_row)


def bag_of_words(s, words):
    bag = [0 for _ in range(len(words))]

    s_words = nltk.word_tokenize(s)
    s_words = [stemmer.stemWord(word.lower()) for word in s_words]

    for se in s_words:
        for i, w in enumerate(words):
            if w == se:
                bag[i] = 1

    return numpy.array(bag)


def chat():

#########################################   arayüz tasarımı  ##################################################
    class App(QMainWindow):

        def __init__(self):
            super().__init__()
            self.title = 'IMDB Öneri Chatbotu'
            self.left = 1520
            self.top = 520
            self.width = 400
            self.height = 500
            self.initUI()
            i = QListWidgetItem('Dr.Bot: Merhaba nasıl yardımcı olabilirim')
            i.setTextAlignment(Qt.AlignLeft)
            i.setBackground(QColor('#beaed4'))

            self.chatbot.addItem(i)

            self.button.clicked.connect(self.chatt)


        def initUI(self):

            self.setWindowTitle(self.title)
            self.setGeometry(self.left, self.top, self.width, self.height)

            self.textbox = QLineEdit(self)
            self.textbox.move(20, 450)
            self.textbox.resize(280, 40)

            self.button = QPushButton('Gönder', self)
            self.button.resize(100, 40)
            self.button.move(300, 450)
            self.button.setStyleSheet("background-color :#32CD32")

            self.chatbot = QListWidget(self)
            self.chatbot.move(20, 20)
            self.chatbot.resize(360, 400)

            self.show()

        def chattextrun(self, text, color, hizalama):
            i = QListWidgetItem(text)
            i.setTextAlignment(hizalama)
            i.setBackground(QColor(color))


            self.chatbot.addItem(i)

#########################################   arayüz tasarımı  ##################################################
        def chatt(self):

            global kategorilist, puanlist, kontrol
            if kontrol == True:
                kategorilist = []
                puanlist = []
                kontrol = False

            self.chattextrun('You: {}     '.format(self.textbox.text()), '#7fc97f', Qt.AlignRight)

            inp = self.textbox.text()
            self.textbox.clear()

            print('Bot yazıyor..')
            results = model.predict(np.asanyarray([bag_of_words(inp, words)]))[0]
            results_index = numpy.argmax(results)
            tag = labels[results_index]

            if results[results_index] > 0.45:

                for tg in data["intents"]:

                    if tg['tag'] == tag:
                        responses = tg['responses']

                if random.choice(responses) == 'Puan aralığınız nedir?(Örneğin 7 puan üstü, 9 puan altı vs.)':
                    self.chattextrun('Dr.Bot: {}'.format(random.choice(responses)), '#beaed4', Qt.AlignLeft)

                    kategorilist.append(inp)
                elif random.choice(responses) == 'Film önerim':
                    puanlist.append(inp)

                    print(kategorilist, puanlist)
                    randomchoicelist = []
                    for run in range(0, 100):

                        puan = imdblistesi[run][2]
                        if puanlist[0].find('üstü') != -1 or puanlist[0].find('üzeri') != -1 or puanlist[0].find(
                                'ustu') != -1 or puanlist[0].find('uzeri') != -1:
                            if puan > int(puanlist[0][0]):
                                kategori = imdblistesi[run][3]
                                if kategori.find(str(kategorilist[0])) != -1 or kategori.find(
                                        str(kategorilist[0]).capitalize()) != -1:
                                    randomchoicelist.append(
                                        'Film önerim: {} \nTürü: {}\nYılı: {} \nPuan: {}\nKonusu: {} '.format(
                                            imdblistesi[run][0], kategori, imdblistesi[run][1], puan,
                                            imdblistesi[run][4]))
                                    kontrol = True

                        if puanlist[0].find('altı') != -1 or puanlist[0].find('alt') != -1:

                            if puan < int(puanlist[0][0]):
                                kategori = imdblistesi[run][3]
                                if kategori.find(str(kategorilist[0])) != -1:
                                    randomchoicelist.append(
                                        'Film önerim: {} \nTürü: {}\nYılı: {} \nPuan: {}\nKonusu: {} '.format(
                                            imdblistesi[run][0], kategori, imdblistesi[run][1], puan,
                                            imdblistesi[run][4]))
                                    kontrol = True
                    if len(randomchoicelist) == 0:
                        self.chattextrun('Dr.Bot: Aradığınız kriterlerde bir film bulunamadı', '#beaed4', Qt.AlignLeft)



                    else:
                        word = random.choice(randomchoicelist).split('Konusu')
############################### filmin konusu bölümündeki kelimeleri chatbot ekranına sığdırmak için yazılmış kod bölümü ###################
                        self.chattextrun('Dr.Bot: {}'.format(word[0]), '#beaed4', Qt.AlignLeft)

                        word = word[1][1:].split(' ')
                        print(len(word))

                        try:
                            for k in range(0, len(word), 5):
                                a = word[k] + ' ' + word[k + 1] + ' ' + word[k + 2] + ' ' + word[k + 3] + ' ' + word[
                                    k + 4] + ' '
                                self.chattextrun('{}'.format(a), '#beaed4', Qt.AlignLeft)

                        except IndexError:

                            if (len(word)) % 5 == 1:
                                self.chattextrun('{}'.format(word[-1]), '#beaed4', Qt.AlignLeft)
                            if (len(word)) % 5 == 2:
                                self.chattextrun('{}'.format(word[-2] + word[-1]), '#beaed4', Qt.AlignLeft)
                            if (len(word)) % 5 == 3:
                                self.chattextrun('{}'.format(word[-3] + word[-2] + word[-1]), '#beaed4', Qt.AlignLeft)
                            if (len(word)) % 5 == 4:
                                self.chattextrun('{}'.format(word[-4] + word[-3] + word[-2] + word[-1]), '#beaed4',
                                                 Qt.AlignLeft)

############################### film önerisi bölümündeki kelimeleri chatbot ekranına sığdırmak için yazılmış kod bölümü ################


                else:
                    self.chattextrun('Dr.Bot: {}'.format(random.choice(responses)), '#beaed4', Qt.AlignLeft)


################################ threshold değerinin altında kalan kelimeler ile karşılaşıldığında çalışacak kod satırı##########

            else:
                self.chattextrun('Dr.Bot: Tam olarak anlayamadım', '#beaed4', Qt.AlignLeft)

    if __name__ == '__main__':
        app = QApplication(sys.argv)
        ex = App()
        sys.exit(app.exec_())


chat()
